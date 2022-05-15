import openpyxl as xl
wb = xl.load_workbook('sudoku.xlsx')

ws = wb['Sheet1 (3)']



row_1 = [1,2,3,4,5,6,7,8,9]
c1_x = []
for i in ws.iter_rows(min_row = 1, min_col= 1, max_col = 9 , max_row = 1, values_only = True):
    c1_x.append(i)
for i in c1_x[0]:
    if i in row_1:
        row_1.remove(i)


row_2 = [1,2,3,4,5,6,7,8,9]
c2_x = []
for i in ws.iter_rows( min_row = 2,min_col = 1, max_col = 9 , max_row = 2, values_only = True):
    c2_x.append(i)
for i in c2_x[0]:
    if i in row_2:
        row_2.remove(i)

row_3 = [1,2,3,4,5,6,7,8,9]
c3_x = []
for i in ws.iter_rows(min_row = 3, min_col= 1, max_col = 9 , max_row = 3, values_only = True):
    c3_x.append(i)
for i in c3_x[0]:
    if i in row_3:
        row_3.remove(i)

row_4 = [1,2,3,4,5,6,7,8,9]
c4_x = []
for i in ws.iter_rows(min_row = 4, min_col= 1, max_col = 9, max_row = 4, values_only = True):
    c4_x.append(i)
for i in c4_x[0]:
    if i in row_4:
        row_4.remove(i)

row_5 = [1,2,3,4,5,6,7,8,9]
c5_x = []
for i in ws.iter_rows(min_row = 5, min_col= 1, max_col = 9, max_row = 5, values_only = True):
    c5_x.append(i)
for i in c5_x[0]:
    if i in row_5:
        row_5.remove(i)

row_6 = [1,2,3,4,5,6,7,8,9]
c6_x = []
for i in ws.iter_rows(min_row = 6, min_col= 1, max_col = 9, max_row = 6, values_only = True):
    c6_x.append(i)
for i in c6_x[0]:
    if i in row_6:
        row_6.remove(i)

row_7 = [1,2,3,4,5,6,7,8,9]
c7_x = []
for i in ws.iter_rows(min_row = 7, min_col= 1, max_col = 9, max_row = 7, values_only = True):
    c7_x.append(i)
for i in c7_x[0]:
    if i in row_7:
        row_7.remove(i)

row_8 = [1,2,3,4,5,6,7,8,9]
c8_x = []
for i in ws.iter_rows(min_row = 8, min_col= 1, max_col = 9, max_row = 8, values_only = True):
    c8_x.append(i)
for i in c8_x[0]:
    if i in row_8:
        row_8.remove(i)

row_9 = [1,2,3,4,5,6,7,8,9]
c9_x = []
for i in ws.iter_rows(min_row = 9, min_col= 1, max_col = 9, max_row = 9, values_only = True):
    c9_x.append(i)
for i in c9_x[0]:
    if i in row_9:
        row_9.remove(i)
