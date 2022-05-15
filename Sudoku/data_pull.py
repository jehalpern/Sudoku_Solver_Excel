import openpyxl as xl
wb = xl.load_workbook('sudoku.xlsx')

ws = wb['Sheet1 (3)']


block_1 = [1,2,3,4,5,6,7,8,9]
b1_x = []
for i in ws.iter_rows(min_row = 1, max_col = 3, max_row = 3, values_only = True):
    b1_x.append(list(i))
for i in b1_x[0]:
    if i in block_1:
        block_1.remove(i)
for i in b1_x[1]:
    if i in block_1:
        block_1.remove(i)
for i in b1_x[2]:
    if i in block_1:
        block_1.remove(i)

block_2 = [1,2,3,4,5,6,7,8,9]
b2_x = []
for i in ws.iter_rows(min_row = 4, max_col = 3, max_row = 6, values_only = True):
    b2_x.append(i)
for i in b2_x[0]:
    if i in block_2:
        block_2.remove(i)
for i in b2_x[1]:
    if i in block_1:
        block_2.remove(i)
for i in b2_x[2]:
    if i in block_2:
        block_2.remove(i)

block_3 = [1,2,3,4,5,6,7,8,9]
b3_x = []
for i in ws.iter_rows(min_row = 7, max_col = 3, max_row = 9, values_only = True):
    b3_x.append(i)
for i in b3_x[0]:
    if i in block_3:
        block_3.remove(i)
for i in b3_x[1]:
    if i in block_3:
        block_3.remove(i)
for i in b3_x[2]:
    if i in block_3:
        block_3.remove(i)

block_4 = [1,2,3,4,5,6,7,8,9]
b4_x = []
for i in ws.iter_cols(min_row = 1, min_col= 4, max_col = 6, max_row = 3, values_only = True):
    b4_x.append(i)
for i in b4_x[0]:
    if i in block_4:
        block_4.remove(i)
for i in b4_x[1]:
    if i in block_4:
        block_4.remove(i)
for i in b4_x[2]:
    if i in block_4:
        block_4.remove(i)

block_5 = [1,2,3,4,5,6,7,8,9]
b5_x = []
for i in ws.iter_cols(min_row = 4, min_col= 4, max_col = 6, max_row = 6, values_only = True):
    b5_x.append(i)
for i in b5_x[0]:
    if i in block_5:
        block_5.remove(i)
for i in b5_x[1]:
    if i in block_5:
        block_5.remove(i)
for i in b5_x[2]:
    if i in block_5:
        block_5.remove(i)

block_6 = [1,2,3,4,5,6,7,8,9]
b6_x = []
for i in ws.iter_cols(min_row = 7, min_col= 4, max_col = 6, max_row = 9, values_only = True):
    b6_x.append(i)
for i in b6_x[0]:
    if i in block_6:
        block_6.remove(i)
for i in b6_x[1]:
    if i in block_6:
        block_6.remove(i)
for i in b6_x[2]:
    if i in block_6:
        block_6.remove(i)

block_7 = [1,2,3,4,5,6,7,8,9]
b7_x = []
for i in ws.iter_cols(min_row = 1, min_col= 7, max_col = 9, max_row = 3, values_only = True):
    b7_x.append(i)
for i in b7_x[0]:
    if i in block_7:
        block_7.remove(i)
for i in b7_x[1]:
    if i in block_7:
        block_7.remove(i)
for i in b7_x[2]:
    if i in block_7:
        block_7.remove(i)

block_8 = [1,2,3,4,5,6,7,8,9]
b8_x = []
for i in ws.iter_cols(min_row = 4, min_col= 7, max_col = 9, max_row = 6, values_only = True):
    b8_x.append(i)
for i in b8_x[0]:
    if i in block_8:
        block_8.remove(i)
for i in b8_x[1]:
    if i in block_8:
        block_8.remove(i)
for i in b8_x[2]:
    if i in block_8:
        block_8.remove(i)

block_9 = [1,2,3,4,5,6,7,8,9]
b9_x = []
for i in ws.iter_cols(min_row = 7, min_col= 7, max_col = 9, max_row = 9, values_only = True):
    b9_x.append(i)
for i in b9_x[0]:
    if i in block_9:
        block_9.remove(i)
for i in b9_x[1]:
    if i in block_9:
        block_9.remove(i)
for i in b9_x[2]:
    if i in block_9:
        block_9.remove(i)


#print('block_1:' , block_1)
#print('block_2:' , block_2)
#print('block_3:' , block_3)
#print('block_4:' , block_4)
#print('block_5:' , block_5)
#print('block_6:' , block_6)
#print('block_7:' , block_7)
#print('block_8:' , block_8)
#print('block_9:' , block_9)
