import openpyxl as xl
wb = xl.load_workbook('sudoku.xlsx')

ws = wb['Sheet1 (3)']



col_1 = [1,2,3,4,5,6,7,8,9]
c1_x = []
for i in ws.iter_cols(min_row = 1, min_col= 0, max_col = 1 , max_row = 9, values_only = True):
    c1_x.append(i)
for i in c1_x[0]:
    if i in col_1:
        col_1.remove(i)

col_2 = [1,2,3,4,5,6,7,8,9]
c2_x = []
for i in ws.iter_cols( min_row = 1,min_col = 2, max_col = 2 , max_row = 9, values_only = True):
    c2_x.append(i)
for i in c2_x[0]:
    if i in col_2:
        col_2.remove(i)

col_3 = [1,2,3,4,5,6,7,8,9]
c3_x = []
for i in ws.iter_cols(min_row = 1, min_col= 3, max_col = 3 , max_row = 9, values_only = True):
    c3_x.append(i)
for i in c3_x[0]:
    if i in col_3:
        col_3.remove(i)

col_4 = [1,2,3,4,5,6,7,8,9]
c4_x = []
for i in ws.iter_cols(min_row = 1, min_col= 4, max_col = 4 , max_row = 9, values_only = True):
    c4_x.append(i)
for i in c4_x[0]:
    if i in col_4:
        col_4.remove(i)

col_5 = [1,2,3,4,5,6,7,8,9]
c5_x = []
for i in ws.iter_cols(min_row = 1, min_col= 5, max_col = 5 , max_row = 9, values_only = True):
    c5_x.append(i)
for i in c5_x[0]:
    if i in col_5:
        col_5.remove(i)

col_6 = [1,2,3,4,5,6,7,8,9]
c6_x = []
for i in ws.iter_cols(min_row = 1, min_col= 6, max_col = 6 , max_row = 9, values_only = True):
    c6_x.append(i)
for i in c6_x[0]:
    if i in col_6:
        col_6.remove(i)

col_7 = [1,2,3,4,5,6,7,8,9]
c7_x = []
for i in ws.iter_cols(min_row = 1, min_col= 7, max_col = 7, max_row = 9, values_only = True):
    c7_x.append(i)
for i in c7_x[0]:
    if i in col_7:
        col_7.remove(i)

col_8 = [1,2,3,4,5,6,7,8,9]
c8_x = []
for i in ws.iter_cols(min_row = 1, min_col= 8, max_col = 8 , max_row = 9, values_only = True):
    c8_x.append(i)
for i in c8_x[0]:
    if i in col_8:
        col_8.remove(i)

col_9 = [1,2,3,4,5,6,7,8,9]
c9_x = []
for i in ws.iter_cols(min_row = 1, min_col= 9, max_col = 9 , max_row = 9, values_only = True):
    c9_x.append(i)
for i in c9_x[0]:
    if i in col_9:
        col_9.remove(i)
