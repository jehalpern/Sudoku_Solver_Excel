import openpyxl as xl
wb = xl.load_workbook('sudoku.xlsx')
from data_pull import block_1, block_2, block_3, block_4, block_5, block_6, block_7, block_8, block_9
from col import col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9
from row import row_1,row_2, row_3, row_4, row_5, row_6, row_7, row_8, row_9

ws = wb['Sheet1 (3)']

cell_coordinates = []
def coordinate(cell):
    if ws[cell].value == None:
        cell_list = (list(cell))
        cell_coordinates.clear()
        if cell_list[0] == 'a':
            cell_coordinates.append(col_1)
        elif cell_list[0] == 'b':
            cell_coordinates.append(col_2)
        elif cell_list[0] == 'c':
            cell_coordinates.append(col_3)
        elif cell_list[0] == 'd':
            cell_coordinates.append(col_4)
        elif cell_list[0] == 'e':
            cell_coordinates.append(col_5)
        elif cell_list[0] == 'f':
            cell_coordinates.append(col_6)
        elif cell_list[0] == 'g':
            cell_coordinates.append(col_7)
        elif cell_list[0] == 'h':
            cell_coordinates.append(col_8)
        elif cell_list[0] == 'i':
            cell_coordinates.append(col_9)

        if cell_list[1] == '1':
            cell_coordinates.append(row_1)
        elif cell_list[1] == '2':
            cell_coordinates.append(row_2)
        elif cell_list[1] == '3':
            cell_coordinates.append(row_3)
        elif cell_list[1] == '4':
            cell_coordinates.append(row_4)
        elif cell_list[1] == '5':
            cell_coordinates.append(row_5)
        elif cell_list[1] == '6':
            cell_coordinates.append(row_6)
        elif cell_list[1] == '7':
            cell_coordinates.append(row_7)
        elif cell_list[1] == '8':
            cell_coordinates.append(row_8)
        elif cell_list[1] == '9':
            cell_coordinates.append(row_9)

        if cell_list[0] == 'a' and cell_list[1] == '1':
            cell_coordinates.append(block_1)
        elif cell_list[0] == 'a' and cell_list[1] == '2':
            cell_coordinates.append(block_1)
        elif cell_list[0] == 'a' and cell_list[1] == '3':
            cell_coordinates.append(block_1)

        elif cell_list[0] == 'b' and cell_list[1] == '1':
            cell_coordinates.append(block_1)
        elif cell_list[0] == 'b' and cell_list[1] == '2':
            cell_coordinates.append(block_1)
        elif cell_list[0] == 'b' and cell_list[1] == '3':
            cell_coordinates.append(block_1)

        elif cell_list[0] == 'c' and cell_list[1] == '1':
            cell_coordinates.append(block_1)
        elif cell_list[0] == 'c' and cell_list[1] == '2':
            cell_coordinates.append(block_1)
        elif cell_list[0] == 'c' and cell_list[1] == '3':
            cell_coordinates.append(block_1)

        elif cell_list[0] == 'a' and cell_list[1] == '4':
            cell_coordinates.append(block_2)
        elif cell_list[0] == 'a' and cell_list[1] == '5':
            cell_coordinates.append(block_2)
        elif cell_list[0] == 'a' and cell_list[1] == '6':
            cell_coordinates.append(block_2)

        elif cell_list[0] == 'b' and cell_list[1] == '4':
            cell_coordinates.append(block_2)
        elif cell_list[0] == 'b' and cell_list[1] == '5':
            cell_coordinates.append(block_2)
        elif cell_list[0] == 'b' and cell_list[1] == '6':
            cell_coordinates.append(block_2)

        elif cell_list[0] == 'c' and cell_list[1] == '4':
            cell_coordinates.append(block_2)
        elif cell_list[0] == 'c' and cell_list[1] == '5':
            cell_coordinates.append(block_2)
        elif cell_list[0] == 'c' and cell_list[1] == '6':
            cell_coordinates.append(block_2)

        elif cell_list[0] == 'a' and cell_list[1] == '7':
            cell_coordinates.append(block_3)
        elif cell_list[0] == 'a' and cell_list[1] == '8':
            cell_coordinates.append(block_3)
        elif cell_list[0] == 'a' and cell_list[1] == '9':
            cell_coordinates.append(block_3)

        elif cell_list[0] == 'b' and cell_list[1] == '7':
            cell_coordinates.append(block_3)
        elif cell_list[0] == 'b' and cell_list[1] == '8':
            cell_coordinates.append(block_3)
        elif cell_list[0] == 'b' and cell_list[1] == '9':
            cell_coordinates.append(block_3)

        elif cell_list[0] == 'c' and cell_list[1] == '7':
            cell_coordinates.append(block_3)
        elif cell_list[0] == 'c' and cell_list[1] == '8':
            cell_coordinates.append(block_3)
        elif cell_list[0] == 'c' and cell_list[1] == '9':
            cell_coordinates.append(block_3)

        elif cell_list[0] == 'd' and cell_list[1] == '1':
            cell_coordinates.append(block_4)
        elif cell_list[0] == 'd' and cell_list[1] == '2':
            cell_coordinates.append(block_4)
        elif cell_list[0] == 'd' and cell_list[1] == '3':
            cell_coordinates.append(block_4)

        elif cell_list[0] == 'e' and cell_list[1] == '1':
            cell_coordinates.append(block_4)
        elif cell_list[0] == 'e' and cell_list[1] == '2':
            cell_coordinates.append(block_4)
        elif cell_list[0] == 'e' and cell_list[1] == '3':
            cell_coordinates.append(block_4)

        elif cell_list[0] == 'f' and cell_list[1] == '1':
            cell_coordinates.append(block_4)
        elif cell_list[0] == 'f' and cell_list[1] == '2':
            cell_coordinates.append(block_4)
        elif cell_list[0] == 'f' and cell_list[1] == '3':
            cell_coordinates.append(block_4)

        elif cell_list[0] == 'd' and cell_list[1] == '4':
            cell_coordinates.append(block_5)
        elif cell_list[0] == 'd' and cell_list[1] == '5':
            cell_coordinates.append(block_5)
        elif cell_list[0] == 'd' and cell_list[1] == '6':
            cell_coordinates.append(block_5)

        elif cell_list[0] == 'e' and cell_list[1] == '4':
            cell_coordinates.append(block_5)
        elif cell_list[0] == 'e' and cell_list[1] == '5':
            cell_coordinates.append(block_5)
        elif cell_list[0] == 'e' and cell_list[1] == '6':
            cell_coordinates.append(block_5)

        elif cell_list[0] == 'f' and cell_list[1] == '4':
            cell_coordinates.append(block_5)
        elif cell_list[0] == 'f' and cell_list[1] == '5':
            cell_coordinates.append(block_5)
        elif cell_list[0] == 'f' and cell_list[1] == '6':
            cell_coordinates.append(block_5)

        elif cell_list[0] == 'd' and cell_list[1] == '7':
            cell_coordinates.append(block_6)
        elif cell_list[0] == 'd' and cell_list[1] == '8':
            cell_coordinates.append(block_6)
        elif cell_list[0] == 'd' and cell_list[1] == '9':
            cell_coordinates.append(block_6)

        elif cell_list[0] == 'e' and cell_list[1] == '7':
            cell_coordinates.append(block_6)
        elif cell_list[0] == 'e' and cell_list[1] == '8':
            cell_coordinates.append(block_6)
        elif cell_list[0] == 'e' and cell_list[1] == '9':
            cell_coordinates.append(block_6)

        elif cell_list[0] == 'f' and cell_list[1] == '7':
            cell_coordinates.append(block_6)
        elif cell_list[0] == 'f' and cell_list[1] == '8':
            cell_coordinates.append(block_6)
        elif cell_list[0] == 'f' and cell_list[1] == '9':
            cell_coordinates.append(block_6)

        elif cell_list[0] == 'g' and cell_list[1] == '1':
            cell_coordinates.append(block_7)
        elif cell_list[0] == 'g' and cell_list[1] == '2':
            cell_coordinates.append(block_7)
        elif cell_list[0] == 'g' and cell_list[1] == '3':
            cell_coordinates.append(block_7)

        elif cell_list[0] == 'h' and cell_list[1] == '1':
            cell_coordinates.append(block_7)
        elif cell_list[0] == 'h' and cell_list[1] == '2':
            cell_coordinates.append(block_7)
        elif cell_list[0] == 'h' and cell_list[1] == '3':
            cell_coordinates.append(block_7)

        elif cell_list[0] == 'i' and cell_list[1] == '1':
            cell_coordinates.append(block_7)
        elif cell_list[0] == 'i' and cell_list[1] == '2':
            cell_coordinates.append(block_7)
        elif cell_list[0] == 'i' and cell_list[1] == '3':
            cell_coordinates.append(block_7)

        elif cell_list[0] == 'g' and cell_list[1] == '4':
            cell_coordinates.append(block_8)
        elif cell_list[0] == 'g' and cell_list[1] == '5':
            cell_coordinates.append(block_8)
        elif cell_list[0] == 'g' and cell_list[1] == '6':
            cell_coordinates.append(block_8)

        elif cell_list[0] == 'h' and cell_list[1] == '4':
            cell_coordinates.append(block_8)
        elif cell_list[0] == 'h' and cell_list[1] == '5':
            cell_coordinates.append(block_8)
        elif cell_list[0] == 'h' and cell_list[1] == '6':
            cell_coordinates.append(block_8)

        elif cell_list[0] == 'i' and cell_list[1] == '4':
            cell_coordinates.append(block_8)
        elif cell_list[0] == 'i' and cell_list[1] == '5':
            cell_coordinates.append(block_8)
        elif cell_list[0] == 'i' and cell_list[1] == '6':
            cell_coordinates.append(block_8)

        elif cell_list[0] == 'g' and cell_list[1] == '7':
            cell_coordinates.append(block_9)
        elif cell_list[0] == 'g' and cell_list[1] == '8':
            cell_coordinates.append(block_9)
        elif cell_list[0] == 'g' and cell_list[1] == '9':
            cell_coordinates.append(block_9)

        elif cell_list[0] == 'h' and cell_list[1] == '7':
            cell_coordinates.append(block_9)
        elif cell_list[0] == 'h' and cell_list[1] == '8':
            cell_coordinates.append(block_9)
        elif cell_list[0] == 'h' and cell_list[1] == '9':
            cell_coordinates.append(block_9)

        elif cell_list[0] == 'i' and cell_list[1] == '7':
            cell_coordinates.append(block_9)
        elif cell_list[0] == 'i' and cell_list[1] == '8':
            cell_coordinates.append(block_9)
        elif cell_list[0] == 'i' and cell_list[1] == '9':
            cell_coordinates.append(block_9)

        samzies = []
        for i in cell_coordinates[0]:
            for j in cell_coordinates[1]:
                for k in cell_coordinates[2]:
                    if i == j == k:
                     samzies.append(i)
        if len(samzies) == 1:
            ws[cell].value = samzies[0]
            wb.save('sudoku.xlsx')











    #elif cell[0] == 'a' and cell[1] == '4' or '5' or '6':
        #cell_coordinates.append(block_2)
    #elif cell[0] == 'b' and cell[1] == '4' or '5' or '6':
        #cell_coordinates.append(block_2)
    #elif cell[0] == 'c' and (cell[1] == '4' or '5' or '6'):
        #cell_coordinates.append(block_2)
